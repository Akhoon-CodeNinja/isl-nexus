import os
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import PyPDFLoader, Docx2txtLoader
from langchain_community.vectorstores import FAISS
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain_core.documents import Document as LangchainDocument

# FAISS index kahan save hoga (root directory mein ek folder ban jayega)
FAISS_INDEX_PATH = "faiss_index"

# Embedding model (Aap yahan OpenAI bhi use kar sakte hain, abhi free local model lagaya hai)
embeddings = HuggingFaceEmbeddings(model_name="all-MiniLM-L6-v2")


def _load_pages(doc_path):
    """Picks the right loader by file extension. Document.FileType now
    allows PDF/DOCX/XLSX (see models.py), so this can no longer assume
    every uploaded file is a PDF the way the original PyPDFLoader-only
    version did — that silently failed (caught by the broad except
    below) for every Word/Excel upload.
    """
    ext = os.path.splitext(doc_path)[1].lower()

    if ext == ".pdf":
        return PyPDFLoader(doc_path).load()

    if ext == ".docx":
        # Requires the `docx2txt` package (pip install docx2txt).
        return Docx2txtLoader(doc_path).load()

    if ext in (".xlsx", ".xls"):
        # A lightweight loader using openpyxl directly rather than pulling
        # in the much heavier `unstructured` package just for this. Each
        # sheet becomes one chunk of tab-separated rows; good enough for
        # the RAG chatbot to search over, not meant to preserve formatting.
        from openpyxl import load_workbook  # pip install openpyxl

        wb = load_workbook(doc_path, data_only=True)
        docs = []
        for sheet in wb.worksheets:
            lines = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    lines.append("\t".join(cells))
            if lines:
                docs.append(
                    LangchainDocument(
                        page_content="\n".join(lines),
                        metadata={"sheet": sheet.title},
                    )
                )
        return docs

    raise ValueError(f"Unsupported file type for indexing: '{ext}'")


def process_and_add_document(doc_path, doc_id, doc_title):
    try:
        # 1. Document ko Load karein (file type ke hisaab se sahi loader)
        pages = _load_pages(doc_path)

        # 2. Text ko chhote Chunks mein todein (taake AI asani se parh sake)
        text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=150)
        chunks = text_splitter.split_documents(pages)
        
        # 3. Har chunk ke sath reference (Document ka naam/ID) attach karein
        for chunk in chunks:
            chunk.metadata['doc_id'] = str(doc_id)
            chunk.metadata['doc_title'] = doc_title
            
        # 4. FAISS index update ya create karein
        if os.path.exists(FAISS_INDEX_PATH):
            # Purana data load karein aur naya add karein (Live Sync)
            db = FAISS.load_local(FAISS_INDEX_PATH, embeddings, allow_dangerous_deserialization=True)
            db.add_documents(chunks)
            db.save_local(FAISS_INDEX_PATH)
        else:
            # Pehli dafa FAISS bana rahe hain
            db = FAISS.from_documents(chunks, embeddings)
            db.save_local(FAISS_INDEX_PATH)
            
        print(f"✅ Document '{doc_title}' successfully indexed in FAISS!")
    except Exception as e:
        print(f"❌ FAISS Indexing failed for '{doc_title}': {e}")